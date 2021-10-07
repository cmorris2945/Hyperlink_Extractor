# importing the needed modules

# os module to list files in directory
import os
# openpyxl module to read excel files
import openpyxl
# textract module to read from doc files
import textract
# PyPDF2 module to read from pdf files
import PyPDF2
# pandas module to write the output to csv file
import pandas as pd
import re

#--------------------------------------------------

# initializing list to store all links from all files
all_urls = []
# failed files list
failed = []

#--------------------------------------------------

# extracts urls from text
def genURLS(text):
    urls = []
    text = text.replace("\\n", " ")

    link_regex = r"(?i)\b((?:https?://|www\d{0,3}[.]|[a-z0-9.\-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|\(([^\s()<>]+|(\([^\s()<>]+\)))*\))+(?:\(([^\s()<>]+|(\([^\s()<>]+\)))*\)|[^\s`!()\[\]{};:'\".,<>?«»“”‘’]))"
    links = re.findall(link_regex, text)

    for link in links:
        urls.append(link[0])

    return urls

#--------------------------------------------------

# a function to append links from different files to the all_urls list
def appendURL(urls):
    for url in urls:
        all_urls.append(url)
    
#--------------------------------------------------

# a function to extract urls from PDF files
def urlPDF(filename):
    # list to store all the urls in the file
    urls = []

    # creating a pdf file object
    pdfFileObj = open(filename, 'rb')
    # creating a pdf reader object
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

    # for loop to iterate over all the pages
    for i in range(pdfReader.numPages):
        # creating a page object
        pageObj = pdfReader.getPage(i)
        # extracting all the urls from this page
        for url in genURLS(pageObj.extractText()):
            # appending the url, filename and the file type
            urls.append([url, filename, "PDF File"])
    
    # closing the pdf file object
    pdfFileObj.close()
    # returning the list of links extracted from the file
    return urls

#--------------------------------------------------

# a function to extract links from a word document
def urlDOC(filename):
    # list to store all the urls in the file
    urls = []

    # using the textract module, we extract the text of the word document as a whole
    text = textract.process(filename)

    # extracting all the links from the whole document
    for url in genURLS(str(text)):
        # appending the url, filename and the file type
        urls.append([url, filename, "Word File"])

    # returning the list of links extracted from the file
    return urls

#--------------------------------------------------

def urlXLS(filename):
    # list to store all the urls in the file
    urls = []

    # using openpyxl, we load the excel file
    wb = openpyxl.load_workbook(filename)
    # and then we open the active sheet (first sheet) in the file
    sheet = wb.active

    # nested loops to iterate over all the cells that are active
    # i iterates over all the active rows
    for i in range(sheet.max_row):
        # j iterates over all the active columns
        for j in range(sheet.max_column):
            # extracting url from every cell [for example: cell(1, 1) is the first cell]
            for url in genURLS(str(sheet.cell(i+1, j+1).value)):
                # appending the url, filename and the file type
                urls.append([url, filename, "Excel File"])

     # returning the list of links extracted from the file
    return urls

#--------------------------------------------------

def extractURLs(file):
    try:
        # defining the extensions that we need
        listOfExtensions = ["doc", "docx", "xls", "xlsx", "pdf"]

        # saving the file name in a variable
        filename = file
        # splitting the file name to get the extension
        file = file.split(".")
        # the extension = the last item after the dot (.)
        extension = file[len(file)-1]
        
        # if the extension of the file in the list of the extensions
        if extension in listOfExtensions:
            # then we check which extension, if it's doc or docx
            if extension == "doc" or extension == "docx":
                # run the function to extract the urls by sending the file name
                urls = urlDOC(filename)
                # then append all the urls that are returned to the all_urls list
                appendURL(urls)
            # if the extension is pdf then...
            elif extension == "pdf":
                # run the function to extract the urls by sending the file name
                urls = urlPDF(filename)
                # then append all the urls that are returned to the all_urls list
                appendURL(urls)
            # if the extension is an excel file then...
            elif extension == "xls" or extension == "xlsx":
                # run the function to extract the urls by sending the file name
                urls = urlXLS(filename)
                # then append all the urls that are returned to the all_urls list
                appendURL(urls)
    except:
        failed.append(filename)

#--------------------------------------------------

# getting a list of files in the current directory
files = os.listdir(os.getcwd())

# iterating over every file and folder and files in folder
for file in files:
    if os.path.isdir(file): 
        path = os.getcwd() + "/" + file
        insidefiles = os.listdir(path)
        for insidefile in insidefiles:
            extractURLs(path + "/" + insidefile)
    else:
        extractURLs(os.getcwd() + "/" + file)

#--------------------------------------------------
"""
At this point we have the list all_urls, which each entry consists of [url, filename, file type]
"""

# we separate each entry from the list to three separate lists
# links list to store the urls
links = []
# filenames list to store the file names
filenames = []
# extensions list to store the extensions
extensions = []

# we iterate over every entry
for url in all_urls:
    # and append the data to every list
    links.append(url[0])
    filenames.append(url[1])
    extensions.append(url[2])

# we define a DataFrame to append it to the csv with three columns
df = pd.DataFrame({
    'Full URLs': links,
    'File Name & Directory': filenames,
    'Extensions' : extensions  
})

failedDF = pd.DataFrame(failed)

# we append the DataFrame to the csv, with indexing enabled
df.to_csv('output.csv', index=True)
# create csv file with failed files
failedDF.to_csv('failed.csv', index=True)